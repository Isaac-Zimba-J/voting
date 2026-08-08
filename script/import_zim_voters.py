#!/usr/bin/env python
"""
One-off bulk voter import for the ZIM association's voter list
(ZIM2.xlsx: columns Full Name, Student Number, Year, Program).

Login credentials: username = Student Number (SIN), password = Student
Number (SIN) - same value for both, per this election's request (unlike
the CSV import feature, which derives the password from a separate NRC
column this file doesn't have).

Usage (run inside the app container, against whichever instance's
database DATABASE_URL points at):
    python script/import_zim_voters.py [path/to/file.xlsx]

Defaults to ZIM2.xlsx in the project root if no path is given. Safe to
re-run: rows whose SIN already exists as a Voter are skipped, not
duplicated.
"""

import os
import sys

import django

CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(CURRENT_DIR)
sys.path.append(PROJECT_ROOT)

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'e_voting.settings')
django.setup()

from openpyxl import load_workbook
from django.db import transaction

from account.models import CustomUser
from voting.models import Voter
from voting.import_utils import split_name


def normalize_sin(raw):
    """Student Number cells come back as int, str, or None depending on
    how the cell was entered/formatted in Excel - normalize to a plain
    digit string with no whitespace, or None if there's nothing usable.
    """
    if raw is None:
        return None
    text = str(raw).strip()
    if text.endswith('.0'):
        text = text[:-2]
    return text or None


def main():
    path = sys.argv[1] if len(sys.argv) > 1 else os.path.join(PROJECT_ROOT, 'ZIM2.xlsx')
    if not os.path.exists(path):
        print(f"File not found: {path}")
        sys.exit(1)

    wb = load_workbook(path, data_only=True)
    ws = wb.active

    created_count = 0
    skipped = []

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        name_cell = row[0] if len(row) > 0 else None
        sin_cell = row[1] if len(row) > 1 else None

        name = (str(name_cell).strip() if name_cell is not None else '')
        sin = normalize_sin(sin_cell)

        if not name and not sin:
            continue  # fully blank row (e.g. a spacer row in the sheet)

        row_label = f"Row {row_num}"
        if not sin:
            skipped.append(f"{row_label} ({name or 'no name'}): missing Student Number")
            continue
        if not name:
            skipped.append(f"{row_label} (SIN {sin}): missing Full Name")
            continue
        if Voter.objects.filter(sin=sin).exists():
            skipped.append(f"{row_label} ({name}): duplicate SIN {sin} - already a voter")
            continue

        first_name, last_name = split_name(name)
        email = f"{sin}@students.local"

        try:
            with transaction.atomic():
                user = CustomUser.objects.create_user(
                    email=email,
                    password=sin,
                    first_name=first_name,
                    last_name=last_name,
                    user_type=2,
                )
                Voter.objects.create(admin=user, sin=sin)
            created_count += 1
        except Exception as e:
            skipped.append(f"{row_label} ({name}, SIN {sin}): could not create account - {e}")

    print(f"Created: {created_count}")
    print(f"Skipped: {len(skipped)}")
    for line in skipped:
        print(f"  - {line}")


if __name__ == '__main__':
    main()
